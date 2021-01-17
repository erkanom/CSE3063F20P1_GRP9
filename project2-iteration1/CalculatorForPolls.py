import shutil

import xlsxwriter
import os
from fuzzywuzzy import fuzz
from Student import Student
from Poll import Poll
from LogAttendance import LogAttendance
from openpyxl import Workbook, load_workbook


class CalculatorForPolls(object):
    def __init__(self, students, dates, pollList):
        self.students = students
        self.dates = dates
        self.pollList = pollList

    def startCalculations(self):
        self.calculateAttendance()
        pollExcell = self.opPoll()
        self.calculatePoll(pollExcell)
        self.closePoll(pollExcell)
        self.analyzePoll()


    def calculateAttendance(self):
        if os.path.exists('Attendance.xlsx'):
            os.remove("./Attendance.xlsx")
        workbook = xlsxwriter.Workbook('Attendance.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, "Student No")
        worksheet.write(0, 1, "First Name")
        worksheet.write(0, 2, "Last Name")
        worksheet.write(0, 3, "Student Information")
        worksheet.write(0, 4, "Attended Courses")
        worksheet.write(0, 5, "Attendance Rate")
        worksheet.write(0, 6, "Attendance Percentage")
        rowIndex = -1
        for student in self.students:
            rowIndex = rowIndex + 1
            if isinstance(student, Student):
                row = []
                row.append(student.getStudentNo())
                row.append(student.getStudentFirstName())
                row.append(student.getStudentLastName())
                row.append(student.getStudentInformation())
                pollList = student.getPollLog()
                attendanceList = student.getAttendanceLogs()
                attendance = 0
                fakeDate = self.dates.copy()
                i = -1
                for date in fakeDate:
                    i = i + 1
                    for poll in pollList:
                        if isinstance(poll, Poll):
                            if poll.getDate() == date:
                                attendance = attendance + 1
                                fakeDate[i] = None
                                break
                i = -1
                for date in fakeDate:
                    i = i + 1
                    for poll in attendanceList:
                        if isinstance(poll, LogAttendance):
                            if poll.getSubmitDate() == date:
                                attendance = attendance + 1
                                fakeDate[i] = None
                                break

            worksheet.write(rowIndex + 1, 0, student.getStudentNo())
            worksheet.write(rowIndex + 1, 1, student.getStudentFirstName())
            worksheet.write(rowIndex + 1, 2, student.getStudentLastName())
            worksheet.write(rowIndex + 1, 3, student.getStudentInformation())
            worksheet.write(rowIndex + 1, 4, attendance)
            worksheet.write(rowIndex + 1, 5, attendance.__str__() + "/" + len(self.dates).__str__())
            worksheet.write(rowIndex + 1, 6, attendance / (len(self.dates)) * 100)

        workbook.close()

    def calculatePoll(self, pollExcell):
        studentRow = 0
        for student in self.students:

            if studentRow == 29:
                a = 0
            row = []

            if isinstance(student, Student):
                column = 4
                row.clear()
                fakePollNames = self.pollList.copy()
                if len(student.getPollLog()) > 0:

                    for poll in student.getPollLog():
                        fakePollNames.remove(poll.getName())
                        row.clear()
                        correctQuestion = 0
                        if isinstance(poll, Poll):
                            studentAnswer = poll.getstudentAnswer()
                            correctAnswer = poll.getanswerList()

                            if len(studentAnswer) > 0:
                                if len(studentAnswer) == len(correctAnswer):
                                    for i in range(len(correctAnswer)):
                                        ratio = fuzz.WRatio(studentAnswer[i], correctAnswer[i])
                                        if ratio > 90:
                                            correctQuestion = correctQuestion + 1
                                            row.append("1")
                                        else:
                                            row.append("0")
                                else:
                                    for i in range(len(correctAnswer)):
                                        row.append("E")
                            else:
                                for i in range(len(correctAnswer)):
                                    row.append("Girmedi")

                        self.appendEcxell(student, row, poll.getName(), pollExcell, studentRow, column, correctQuestion)
                    for fake in range(len(fakePollNames)):
                        pn = fakePollNames[fake]
                        row.clear()
                        row.append("GIRMEDI")
                        if pn != None:
                            self.appendEcxell(student, row, pn, pollExcell, studentRow, column, 0)
                        fakePollNames[fake] = None

                for fake in range(len(fakePollNames)):
                    pn = fakePollNames[fake]
                    row.clear()
                    row.append("GIRMEDI")
                    if pn != None:
                        self.appendEcxell(student, row, pn, pollExcell, studentRow, column, 0)
                    fakePollNames[fake] = None
            studentRow = studentRow + 1

    def analyzePoll(self):
        for pollName in self.pollList:

            questions = []
            distribution = []
            gir = 1
            studentCount = 0
            uniqueAns = []
            uniqueAns.clear()
            ilkforIcingir = 1

            for answerIcinStudent in self.students:
                if isinstance(answerIcinStudent, Student):
                    for answerIcinPoll in answerIcinStudent.getPollLog():
                        if isinstance(answerIcinPoll, Poll):
                            if answerIcinPoll.getName() == pollName:
                                if ilkforIcingir == 1:
                                    for x in range(len(answerIcinPoll.getQuestion())):
                                        uniqueAns.append([])
                                        ilkforIcingir = 0
                                verilenAnswer = answerIcinPoll.getstudentAnswer()
                                for i in range(len(verilenAnswer)):
                                    if uniqueAns[i].count(verilenAnswer[i]) == 0:
                                        uniqueAns[i].append(verilenAnswer[i])

            grap = []
            allAns = []
            for x in uniqueAns:
                if isinstance(x, list):
                    allAns.append(x[:])
                    grap.append(x[:])

            for x in allAns:
                for i in range(len(x)):
                    x[i] = 0

            for student in self.students:
                if isinstance(student, Student):
                    for poll in student.getPollLog():
                        if isinstance(poll, Poll):
                            if poll.getName() == pollName:

                                questions = poll.getQuestion()
                                studentAnswer = poll.getstudentAnswer()
                                correctAnswer = poll.getanswerList()
                                if gir == 1:
                                    for i in range(0, len(correctAnswer)):
                                        distribution.append(0)
                                    gir = 0
                                if len(studentAnswer) == len(correctAnswer):
                                    studentCount = studentCount + 1
                                    for qindex in range(len(poll.getanswerList())):

                                        ratio = fuzz.WRatio(studentAnswer[qindex], correctAnswer[qindex])
                                        if ratio > 90:
                                            distribution[qindex] = distribution[qindex] + 1
                                        for ansIn in range(len(grap[qindex])):
                                            ratioF = fuzz.WRatio(studentAnswer[qindex], grap[qindex][ansIn])
                                            if ratioF == 100:
                                                allAns[qindex][ansIn] = allAns[qindex][ansIn] + 1

            self.graphDisto(pollName, questions, distribution, studentCount, uniqueAns, allAns)

    def graphDisto(self, pollName, questions, distribution, studentCount, uniqueAns, allAns):

        workbook = xlsxwriter.Workbook(pollName + 'Pie_Qdistro.xlsx')
        pieS = workbook.add_worksheet("Attendance")

        chart2 = workbook.add_chart({'type': 'pie'})
        pieS.write(0, 0, "Attanded")
        pieS.write(1, 0, "Not Attanded")
        pieS.write(0, 1, studentCount)
        pieS.write(1, 1, len(self.students) - studentCount)

        chart2.add_series({
            'name': 'Pie sales data',
            'categories': '=Attendance!$A$1:$A$2',
            'values': '=Attendance!$B$1:$B$2',
            'points': [
                {'fill': {'color': '#5ABA10'}},
                {'fill': {'color': '#FE110E'}},

            ],
        })
        chart2.set_title({'name': 'Attendance'})
        pieS.insert_chart('B10', chart2, {'x_offset': 10, 'y_offset': 10})

        for x in range(len(questions)):
            distro = workbook.add_worksheet("Q" + (x + 1).__str__())
            a = []
            a.clear()
            questioncount = 0
            for i in range(len(allAns[x])):
                a.append("A" + i.__str__())
                questioncount = questioncount + 1
            correctIndex = 0
            for i in range(len(distribution)):
                for j in range(len(allAns[x])):
                    if allAns[x][j] == distribution[i]:
                        correctIndex = j

            temp0 = allAns[x][0]
            allAns[x][0] = allAns[x][correctIndex]
            allAns[x][correctIndex] = temp0

            data = [
                a,
                allAns[x],
                uniqueAns[x]
            ]
            distro.write_column('B2', data[0])
            distro.write_column('A2', data[1])
            distro.write_column('C2', data[2])
            chart1 = workbook.add_chart({'type': 'bar'})
            chart1.add_series({
                'categories': '=' + 'Q' + (x + 1).__str__() + '!$B$2:$B$' + (len(allAns[x]) + 1).__str__(),
                'values': '=' + 'Q' + (x + 1).__str__() + '!$A$2:$A$' + (questioncount + 1).__str__(),
                'points': [
                    {'fill': {'color': '#5ABA10'}},

                ],
            })
            chart1.set_style(11)
            chart1.set_title({'name': questions[x]})

            distro.insert_chart('D20', chart1, {'x_offset': 25, 'y_offset': 10})


        workbook.close()

    def opPoll(self):
        pollExcell = []
        for poll in self.pollList:
            if os.path.exists(poll):
                os.remove(poll)

            workbook = xlsxwriter.Workbook(poll + ".xlsx")

            workbook.add_worksheet("sheet1")
            pollExcell.append(workbook)

        return pollExcell

    def closePoll(self, pollExcell):
        for x in pollExcell:
            x.close()

    def appendEcxell(self, student, row, pollName, pollExcell, studentRow, column, corA):


        workbook = load_workbook(filename="Global_LOG.xlsx")
        sheet = workbook.active
        rc=column+2
        if studentRow>0:
            for i in range(1,500):
                if sheet.cell(row=studentRow+1,column=rc+i).value!=None:
                    rc=rc+1
                    break

        print(column,rc)
        sheet.cell(row=studentRow+1, column=1).value=student.getStudentNo()
        sheet.cell(row=studentRow+1, column=2).value =student.getStudentFirstName()
        sheet.cell(row=studentRow+1, column=3).value =student.getStudentLastName()
        sheet.cell(row=studentRow+1, column=4).value =student.getStudentInformation()
        sheet.cell(row=studentRow + 1, column=5).value = pollName
        if len(row)>0:
            sheet.cell(row=studentRow+1,column=rc).value=len(row)
        else:
            print("G")
        sheet.cell(row=studentRow+1,column=rc+1).value=(corA / len(row)) * 100
        workbook.save(filename="Global_LOG.xlsx")


        pollName = pollName + ".xlsx"
        for x in pollExcell:
            if x.filename == pollName:
                sheet1 = x.get_worksheet_by_name("sheet1")
                sheet1.write(studentRow, 0, student.getStudentNo())
                sheet1.write(studentRow, 1, student.getStudentFirstName())
                sheet1.write(studentRow, 2, student.getStudentLastName())
                sheet1.write(studentRow, 3, student.getStudentInformation())

                for x in range(len(row)):
                    sheet1.write(studentRow, column + x, row[x])
                if row[0] != "GIRMEDI":
                    sheet1.write(studentRow, (column + x + 1), len(row))
                    sheet1.write(studentRow, (column + x + 2), corA.__str__() + "/" + len(row).__str__())
                    sheet1.write(studentRow, (column + x + 3), (corA / len(row)) * 100)
